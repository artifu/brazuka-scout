"""
import_brazukas1.py
Imports historical football data from brazukas1.xlsx into Supabase.
Covers Jul 2021 – Jul 2023 data for Brazuka US (team_id=1) and Receba FC (team_id=2).

Sheets used:
  - GamesId: 122 games
  - EventID: Spring 2023 goals/assists
  - PlayerID: player lookup
  - TeamsID: team name lookup
"""

import os
import sys
import datetime
from pathlib import Path

import openpyxl
from supabase import create_client

# ---------------------------------------------------------------------------
# Load environment
# ---------------------------------------------------------------------------
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    sys.exit("ERROR: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# ---------------------------------------------------------------------------
# Player normalizer
# ---------------------------------------------------------------------------
from player_normalizer import PlayerNormalizer
norm = PlayerNormalizer()

# ---------------------------------------------------------------------------
# Load spreadsheet
# ---------------------------------------------------------------------------
XLSX = Path(__file__).parent / "spreadsheet data" / "brazukas1.xlsx"
wb = openpyxl.load_workbook(str(XLSX), data_only=True)

# ---------------------------------------------------------------------------
# Build lookup tables from spreadsheet
# ---------------------------------------------------------------------------

def load_teams_lookup(wb):
    """Returns dict {team_id_int: team_name_str}"""
    ws = wb["TeamsID"]
    lookup = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or row[0] == "TeamID":
            continue
        tid = int(row[0])
        name = str(row[1]).strip() if row[1] else ""
        lookup[tid] = name
    return lookup


def load_players_lookup(wb):
    """Returns dict {player_id_int: player_name_str}"""
    ws = wb["PlayerID"]
    lookup = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or row[0] == "PlayerID":
            continue
        try:
            pid = int(row[0])
        except (TypeError, ValueError):
            continue
        name = str(row[1]).strip() if row[1] else ""
        lookup[pid] = name
    return lookup


teams_lookup = load_teams_lookup(wb)
players_lookup = load_players_lookup(wb)

# Our team IDs in the spreadsheet
BRAZUKA_TEAM_ID = 1   # "Brazukas"
RECEBA_TEAM_ID = 2    # "RECEBA FC"
OUR_TEAM_IDS = {BRAZUKA_TEAM_ID, RECEBA_TEAM_ID}

# Supabase team_id values
SUPABASE_BRAZUKA_TEAM_ID = 1
SUPABASE_RECEBA_TEAM_ID = 2

# ---------------------------------------------------------------------------
# Helper: determine result from our team's POV
# ---------------------------------------------------------------------------
RESULT_MAP = {
    "WIN": "W",
    "DEFEAT": "L",
    "TIE": "D",
}


def our_result(home_result, we_are_home):
    """Convert spreadsheet 'Home Result' to our team's W/L/D."""
    if not home_result:
        return None
    hr = home_result.strip().upper()
    if we_are_home:
        return RESULT_MAP.get(hr)
    # Flip
    flip = {"WIN": "DEFEAT", "DEFEAT": "WIN", "TIE": "TIE"}
    return RESULT_MAP.get(flip.get(hr, hr))


# ---------------------------------------------------------------------------
# Org/Venue → venue / field mapping
# ---------------------------------------------------------------------------
ORG_TO_VENUE = {
    "Magnuson": "Magnuson Park",
    "Redmond": "Redmond",
    "SODO": "SODO",
}

VENUE_TO_FIELD = {
    "LAX": "LAX",
    "Side": "Side",
    "Main": "Main",
    "Issaquah": "Issaquah",
}


# ---------------------------------------------------------------------------
# Season handling
# ---------------------------------------------------------------------------
def get_or_create_season(league_name, team_id, start_date, end_date):
    """
    Look up existing season by (name, team_id) or insert a new one.
    Returns season_id.
    """
    resp = (
        sb.table("seasons")
        .select("id")
        .eq("name", league_name)
        .eq("team_id", team_id)
        .execute()
    )
    if resp.data:
        return resp.data[0]["id"]
    # Insert
    insert_resp = (
        sb.table("seasons")
        .insert({
            "name": league_name,
            "team_id": team_id,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
        })
        .execute()
    )
    return insert_resp.data[0]["id"]


# ---------------------------------------------------------------------------
# Step 1: Collect all games from GamesId, grouped by league+our_team_id
# ---------------------------------------------------------------------------
ws_games = wb["GamesId"]
games_rows = list(ws_games.iter_rows(values_only=True))
header = games_rows[0]
# Header: GameID, Home_Team, Home_Score, Away_Team, Away_Score, Date, Time, Org, Venue, League, HomeID, AwayID, Home Result

# Filter only rows where one of the teams is ours
our_games = []
for row in games_rows[1:]:
    if row[0] is None:
        continue
    home_id = int(row[10]) if row[10] is not None else None
    away_id = int(row[11]) if row[11] is not None else None
    if home_id not in OUR_TEAM_IDS and away_id not in OUR_TEAM_IDS:
        continue
    # Skip games with null scores
    if row[2] is None and row[4] is None:
        continue
    our_games.append(row)

# Group by (league, our_team_id) to find season date ranges
from collections import defaultdict
season_dates = defaultdict(list)
for row in our_games:
    game_date = row[5]
    home_id = int(row[10]) if row[10] is not None else None
    away_id = int(row[11]) if row[11] is not None else None
    league = str(row[9]).strip() if row[9] else "Unknown"

    if home_id in OUR_TEAM_IDS:
        our_tid = home_id
    else:
        our_tid = away_id

    supabase_tid = SUPABASE_BRAZUKA_TEAM_ID if our_tid == BRAZUKA_TEAM_ID else SUPABASE_RECEBA_TEAM_ID
    key = (league, supabase_tid)
    if game_date:
        season_dates[key].append(game_date)

# ---------------------------------------------------------------------------
# Step 2: Create seasons in Supabase
# ---------------------------------------------------------------------------
print("=== Creating seasons ===")
season_id_map = {}  # (league, supabase_tid) -> season_id

for (league, supabase_tid), dates in sorted(season_dates.items()):
    dates_sorted = sorted([d for d in dates if d])
    start_dt = dates_sorted[0].date() if dates_sorted else None
    end_dt = dates_sorted[-1].date() if dates_sorted else None
    sid = get_or_create_season(league, supabase_tid, start_dt, end_dt)
    season_id_map[(league, supabase_tid)] = sid
    print(f"  Season '{league}' team_id={supabase_tid}: id={sid} ({start_dt} → {end_dt})")

# ---------------------------------------------------------------------------
# Step 3: Insert games
# ---------------------------------------------------------------------------
print("\n=== Inserting games ===")
games_inserted = 0
games_skipped = 0
player_warnings = []
# game_id_map: spreadsheet GameID -> supabase game row id
game_id_map = {}

for row in our_games:
    game_id_xl = int(row[0])
    home_team_name = str(row[1]).strip() if row[1] else ""
    home_score = int(row[2]) if row[2] is not None else None
    away_team_name = str(row[3]).strip() if row[3] else ""
    away_score = int(row[4]) if row[4] is not None else None
    game_date = row[5]  # datetime
    org = str(row[7]).strip() if row[7] else None
    venue_sub = str(row[8]).strip() if row[8] else None
    league = str(row[9]).strip() if row[9] else "Unknown"
    home_id = int(row[10]) if row[10] is not None else None
    away_id = int(row[11]) if row[11] is not None else None
    home_result_raw = str(row[12]).strip() if row[12] else None

    if game_date is None:
        continue

    game_date_obj = game_date.date() if isinstance(game_date, datetime.datetime) else game_date
    game_date_str = game_date_obj.isoformat()

    # Which of our teams is in this game?
    if home_id in OUR_TEAM_IDS:
        our_xl_tid = home_id
        we_are_home = True
        opponent_name = away_team_name
        score_us = home_score
        score_opp = away_score
    else:
        our_xl_tid = away_id
        we_are_home = False
        opponent_name = home_team_name
        score_us = away_score
        score_opp = home_score

    supabase_tid = SUPABASE_BRAZUKA_TEAM_ID if our_xl_tid == BRAZUKA_TEAM_ID else SUPABASE_RECEBA_TEAM_ID
    home_or_away = "home" if we_are_home else "away"
    result = our_result(home_result_raw, we_are_home)
    venue_name = ORG_TO_VENUE.get(org, org)
    field_name = VENUE_TO_FIELD.get(venue_sub, venue_sub)
    season_id = season_id_map.get((league, supabase_tid))

    # Check idempotency: does this game already exist?
    existing = (
        sb.table("games")
        .select("id")
        .eq("game_date", game_date_str)
        .eq("team_id", supabase_tid)
        .execute()
    )
    if existing.data:
        game_id_map[game_id_xl] = existing.data[0]["id"]
        print(f"  SKIP (already exists): {game_date_str} vs {opponent_name} (team_id={supabase_tid})")
        games_skipped += 1
        continue

    insert_resp = (
        sb.table("games")
        .insert({
            "game_date": game_date_str,
            "opponent": opponent_name,
            "home_or_away": home_or_away,
            "result": result,
            "score_brazuka": score_us,
            "score_opponent": score_opp,
            "scorers_known": False,
            "team_id": supabase_tid,
            "season_id": season_id,
            "venue": venue_name,
            "field": field_name,
        })
        .execute()
    )
    new_id = insert_resp.data[0]["id"]
    game_id_map[game_id_xl] = new_id
    print(f"  Inserting game {game_date_str} vs {opponent_name} (team_id={supabase_tid})... OK (id={new_id})")
    games_inserted += 1

# ---------------------------------------------------------------------------
# Step 4: Insert goals (and assists) from EventID
# ---------------------------------------------------------------------------
print("\n=== Inserting goals/assists from EventID ===")
ws_events = wb["EventID"]
events_rows = list(ws_events.iter_rows(values_only=True))
# Header: EventID, EventType, Event, TeamID, MatchID, PlayerID, PlayerName, Qty, Desc, Team Name, Date, Temporada

# Collect all Gol rows first, then overlay Assist rows
# Since assists in this sheet are per-player (not per-goal), we store them as notes on goals

# Group gol rows by (match_id, player_id) and assist rows by match_id
from collections import defaultdict
goals_by_match = defaultdict(list)   # match_id -> list of goal dicts
assists_by_match = defaultdict(list)  # match_id -> list of assist dicts

for row in events_rows[1:]:
    if not row or row[1] != "MatchStat":
        continue
    event_type = str(row[2]).strip() if row[2] else ""
    if event_type not in ("Gol", "Assist"):
        continue

    match_id_xl = int(row[4]) if row[4] is not None else None
    player_id_xl = int(row[5]) if row[5] is not None else None
    player_name = str(row[6]).strip() if row[6] else ""
    qty = int(row[7]) if row[7] is not None else 1
    desc = str(row[8]).strip() if row[8] else None

    if match_id_xl is None:
        continue

    if event_type == "Gol":
        # Resolve player
        pid, canonical = norm.resolve_or_flag(player_name)
        if pid is None:
            # Try by spreadsheet PlayerID lookup
            if player_id_xl and player_id_xl in players_lookup:
                pid, canonical = norm.resolve_or_flag(players_lookup[player_id_xl])
        if pid is None:
            warning = f"Player not resolved: '{player_name}' (spreadsheet PlayerID={player_id_xl}) in MatchID={match_id_xl}"
            player_warnings.append(warning)
            print(f"  WARNING: {warning}")

        goals_by_match[match_id_xl].append({
            "player": player_name,
            "player_id": pid,
            "count": qty,
            "notes": desc,
            "assist_player": None,
            "assist_player_id": None,
            "xl_player_id": player_id_xl,
        })

    elif event_type == "Assist":
        pid, canonical = norm.resolve_or_flag(player_name)
        if pid is None:
            if player_id_xl and player_id_xl in players_lookup:
                pid, canonical = norm.resolve_or_flag(players_lookup[player_id_xl])
        assists_by_match[match_id_xl].append({
            "player": player_name,
            "player_id": pid,
            "qty": qty,
            "xl_player_id": player_id_xl,
        })

# Now insert goals and try to match assists
goals_inserted = 0
for match_id_xl, goal_list in goals_by_match.items():
    supabase_game_id = game_id_map.get(match_id_xl)
    if supabase_game_id is None:
        print(f"  WARNING: MatchID={match_id_xl} not found in imported games — skipping goals")
        continue

    # Get assists for this match
    assist_list = assists_by_match.get(match_id_xl, [])

    for goal in goal_list:
        # Try to find a matching assist (same player assisting someone else)
        # The assist data in this sheet is not one-to-one matched, so just attach
        # the first unmatched assist (or leave None)
        assist_player = None
        assist_player_id = None
        if assist_list:
            # Heuristic: if there's only one assist row, attach it
            # Otherwise, we can't reliably match - just leave None
            # The desc field often has info like "Assist do Sergio"
            if goal["notes"]:
                notes_lower = goal["notes"].lower()
                for a in assist_list:
                    if a["player"].lower() in notes_lower:
                        assist_player = a["player"]
                        assist_player_id = a["player_id"]
                        break

        # Check if goal already exists for this game+player
        existing_goal = (
            sb.table("goals")
            .select("id")
            .eq("game_id", supabase_game_id)
            .eq("player", goal["player"])
            .execute()
        )
        if existing_goal.data:
            continue

        sb.table("goals").insert({
            "game_id": supabase_game_id,
            "player": goal["player"],
            "player_id": goal["player_id"],
            "count": goal["count"],
            "notes": goal["notes"],
            "assist_player": assist_player,
            "assist_player_id": assist_player_id,
        }).execute()
        goals_inserted += 1

    # Mark game scorers_known=True
    sb.table("games").update({"scorers_known": True}).eq("id", supabase_game_id).execute()

print(f"  Goals inserted: {goals_inserted}")

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
print("\n=== Summary ===")
print(f"  Games inserted : {games_inserted}")
print(f"  Games skipped  : {games_skipped}")
print(f"  Goals inserted : {goals_inserted}")
print(f"  Player warnings: {len(player_warnings)}")
for w in player_warnings:
    print(f"    WARNING: {w}")

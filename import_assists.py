"""
import_assists.py
Imports assist data into the `assists` table from both spreadsheets.

Sources:
  1. brazukas1.xlsx  — EventID sheet (per-game assist events, Spring 2023)
  2. brazukareceba1.xlsx — per-season sheets (seasonal totals, PLAYER/GOALS/ASSISTS/CARDS table)

Run after the tables and seasons already exist:
  python3 import_assists.py
"""

import os
import sys
import datetime
import re
from pathlib import Path
from collections import defaultdict

import openpyxl
from supabase import create_client

# ── env ───────────────────────────────────────────────────────────────────────
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")
if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    sys.exit("ERROR: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

from player_normalizer import PlayerNormalizer
norm = PlayerNormalizer()

# ── helpers ───────────────────────────────────────────────────────────────────

def find_game_id(game_date_str: str, team_id: int):
    r = sb.table("games").select("id").eq("game_date", game_date_str).eq("team_id", team_id).execute()
    return r.data[0]["id"] if r.data else None


def find_season_id(team_id: int, start_date: datetime.date):
    """Find closest season by start_date within ±45 days."""
    lo = (start_date - datetime.timedelta(days=45)).isoformat()
    hi = (start_date + datetime.timedelta(days=45)).isoformat()
    r = (sb.table("seasons").select("id, start_date")
         .eq("team_id", team_id)
         .gte("start_date", lo).lte("start_date", hi)
         .order("start_date").execute())
    if r.data:
        return r.data[0]["id"]
    return None


def scored_by_from_desc(desc):
    """Try to detect the scorer's name in a description string."""
    if not desc:
        return None, None
    # Search for any known player alias in the description
    desc_lower = desc.lower()
    # Try longer aliases first to avoid partial matches
    best = None
    for alias, player in sorted(norm._index.items(), key=lambda x: -len(x[0])):
        if alias in desc_lower:
            best = player
            break
    if best:
        return best["id"], best["canonical_name"]
    return None, None


def insert_assist(game_id, season_id, team_id, player, player_id, count, scored_by, scored_by_player_id, notes):
    sb.table("assists").insert({
        "game_id": game_id,
        "season_id": season_id,
        "team_id": team_id,
        "player": player,
        "player_id": player_id,
        "count": count,
        "scored_by": scored_by,
        "scored_by_player_id": scored_by_player_id,
        "notes": notes,
    }).execute()


# ── Source 1: brazukas1.xlsx EventID ─────────────────────────────────────────
print("=" * 60)
print("Source 1: brazukas1.xlsx — EventID assists (Spring 2023)")

BRAZUKAS1 = Path(__file__).parent / "spreadsheet data" / "brazukas1.xlsx"
wb1 = openpyxl.load_workbook(str(BRAZUKAS1), data_only=True)

# Build PlayerID → name lookup
players_lookup = {}
for row in wb1["PlayerID"].iter_rows(values_only=True):
    if row[0] is None or row[0] == "PlayerID":
        continue
    try:
        pid = int(row[0])
        players_lookup[pid] = str(row[1]).strip()
    except (TypeError, ValueError):
        pass

# TeamID map (spreadsheet) → Supabase team_id
TEAM_ID_MAP = {1: 1, 2: 2}  # Brazukas→1, RECEBA→2

# Parse EventID rows:  EventID,EventType,Event,TeamID,MatchID,PlayerID,PlayerName,Qty,Desc,TeamName,Date,Season
ws_events = wb1["EventID"]
rows = list(ws_events.iter_rows(values_only=True))

# Build: match_id → list of goal rows (to help derive scored_by)
goals_by_match = defaultdict(list)
assists_rows = []

for row in rows[1:]:
    if not row or row[1] != "MatchStat":
        continue
    event_type = str(row[2]).strip() if row[2] else ""
    match_id = int(row[4]) if row[4] is not None else None
    player_id_xl = int(row[5]) if row[5] is not None else None
    player_name = str(row[6]).strip() if row[6] else ""
    qty = int(row[7]) if row[7] is not None else 1
    desc = str(row[8]).strip() if row[8] else None
    team_id_xl = int(row[3]) if row[3] is not None else None
    game_date = row[10].date() if isinstance(row[10], datetime.datetime) else row[10]

    if match_id is None:
        continue

    if event_type == "Gol":
        pid, canonical = norm.resolve_or_flag(player_name)
        if pid is None and player_id_xl and player_id_xl in players_lookup:
            pid, canonical = norm.resolve_or_flag(players_lookup[player_id_xl])
        goals_by_match[match_id].append({
            "player": canonical,
            "player_id": pid,
            "desc": desc,
            "team_id_xl": team_id_xl,
            "game_date": game_date,
        })
    elif event_type == "Assist":
        pid, canonical = norm.resolve_or_flag(player_name)
        if pid is None and player_id_xl and player_id_xl in players_lookup:
            pid, canonical = norm.resolve_or_flag(players_lookup[player_id_xl])
        assists_rows.append({
            "match_id": match_id,
            "player": canonical,
            "player_id": pid,
            "qty": qty,
            "desc": desc,
            "team_id_xl": team_id_xl,
            "game_date": game_date,
        })

inserted = 0
skipped = 0
warnings = []

for a in assists_rows:
    supabase_team_id = TEAM_ID_MAP.get(a["team_id_xl"])
    if supabase_team_id is None:
        warnings.append(f"Unknown team_id_xl={a['team_id_xl']} for assist by {a['player']}")
        continue

    game_date_str = a["game_date"].isoformat() if a["game_date"] else None
    if not game_date_str:
        warnings.append(f"No date for assist by {a['player']} in match {a['match_id']}")
        continue

    game_id = find_game_id(game_date_str, supabase_team_id)
    if game_id is None:
        warnings.append(f"Game not found: {game_date_str} team_id={supabase_team_id} (assist by {a['player']})")
        continue

    # Idempotency: skip if already exists for this game+player
    existing = (sb.table("assists").select("id")
                .eq("game_id", game_id)
                .eq("player", a["player"])
                .execute())
    if existing.data:
        skipped += 1
        continue

    # Try to determine scored_by:
    # If Qty==1, scan the goal rows for this match to find whose desc mentions the assister
    scored_by_name = None
    scored_by_pid = None

    if a["qty"] == 1:
        # Strategy 1: look for a goal row whose desc contains the assister's name
        for g in goals_by_match.get(a["match_id"], []):
            if g["desc"] and a["player"].split()[0].lower() in g["desc"].lower():
                scored_by_name = g["player"]
                scored_by_pid = g["player_id"]
                break
        # Strategy 2: look for a player name in the assist desc
        if not scored_by_name:
            scored_by_pid, scored_by_name = scored_by_from_desc(a["desc"])
            # Don't set scored_by to the assister themselves
            if scored_by_pid == a["player_id"]:
                scored_by_pid, scored_by_name = None, None

    insert_assist(
        game_id=game_id,
        season_id=None,
        team_id=supabase_team_id,
        player=a["player"],
        player_id=a["player_id"],
        count=a["qty"],
        scored_by=scored_by_name,
        scored_by_player_id=scored_by_pid,
        notes=a["desc"],
    )
    inserted += 1
    sb_label = scored_by_name or "?"
    print(f"  + assist: {a['player']} → {sb_label} ({game_date_str}, count={a['qty']})")

print(f"  Inserted: {inserted}  |  Skipped: {skipped}  |  Warnings: {len(warnings)}")
for w in warnings:
    print(f"  WARNING: {w}")


# ── Source 2: brazukareceba1.xlsx per-season PLAYER tables ───────────────────
print()
print("=" * 60)
print("Source 2: brazukareceba1.xlsx — per-season PLAYER/GOALS/ASSISTS tables")

BRAZUKARECEBA1 = Path(__file__).parent / "spreadsheet data" / "brazukareceba1.xlsx"
wb2 = openpyxl.load_workbook(str(BRAZUKARECEBA1), data_only=True)

# Determine team_id from sheet name
def team_id_from_sheet(sheet_name: str) -> int:
    return 2 if "receba" in sheet_name.lower() else 1


def parse_season_sheets(wb):
    """
    Yields dicts: {team_id, start_date, players: [{name, goals, assists, cards}]}
    for each sheet that has a PLAYER/GOALS/ASSISTS table.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))

        # Find start_date (row 5 or 6, col B)
        start_date = None
        for r in rows[:8]:
            if r[0] in ("Start Date", "start date") and r[1]:
                val = r[1]
                if isinstance(val, datetime.datetime):
                    start_date = val.date()
                break

        if not start_date:
            continue  # can't link to a season without a date

        # Find the header row with PLAYER and ASSISTS columns
        header_row_idx = None
        goals_col = None
        assists_col = None
        for i, row in enumerate(rows):
            row_vals = [str(c).strip().upper() if c else "" for c in row]
            if "PLAYER" in row_vals and "ASSISTS" in row_vals:
                header_row_idx = i
                goals_col = next((j for j, v in enumerate(row_vals) if v == "GOALS"), None)
                assists_col = next((j for j, v in enumerate(row_vals) if v == "ASSISTS"), None)
                player_col = next((j for j, v in enumerate(row_vals) if v == "PLAYER"), None)
                break

        if header_row_idx is None or assists_col is None:
            continue  # no assists data in this sheet

        # Read player rows until we run out of names
        players = []
        for row in rows[header_row_idx + 1:]:
            name = row[player_col] if row[player_col] else None
            if not name or str(name).strip().upper() in ("OTHER", "PLAYER", ""):
                continue
            name = str(name).strip()
            assists = row[assists_col]
            if assists is None or assists == "" or assists == 0:
                continue
            try:
                assists_int = int(assists)
            except (TypeError, ValueError):
                continue
            if assists_int <= 0:
                continue

            goals = None
            if goals_col is not None:
                try:
                    goals = int(row[goals_col]) if row[goals_col] else None
                except (TypeError, ValueError):
                    pass

            players.append({"name": name, "goals": goals, "assists": assists_int})

        if players:
            yield {
                "sheet": sheet_name,
                "team_id": team_id_from_sheet(sheet_name),
                "start_date": start_date,
                "players": players,
            }


inserted2 = 0
skipped2 = 0
warnings2 = []

for block in parse_season_sheets(wb2):
    team_id = block["team_id"]
    season_id = find_season_id(team_id, block["start_date"])

    if season_id is None:
        warnings2.append(f"Season not found for {block['sheet']} (start={block['start_date']}, team_id={team_id})")
        continue

    print(f"\n  Sheet: {block['sheet']}  →  season_id={season_id}  team_id={team_id}")

    for p in block["players"]:
        pid, canonical = norm.resolve_or_flag(p["name"])
        if pid is None:
            warnings2.append(f"  Player not resolved: '{p['name']}' in sheet '{block['sheet']}'")

        # Idempotency: skip if a season-level assist already exists for this player+season
        existing = (sb.table("assists").select("id")
                    .is_("game_id", "null")
                    .eq("season_id", season_id)
                    .eq("player", canonical)
                    .execute())
        if existing.data:
            skipped2 += 1
            continue

        insert_assist(
            game_id=None,
            season_id=season_id,
            team_id=team_id,
            player=canonical,
            player_id=pid,
            count=p["assists"],
            scored_by=None,
            scored_by_player_id=None,
            notes=f"Season total from spreadsheet (goals={p['goals']})",
        )
        inserted2 += 1
        print(f"    + {canonical}: {p['assists']} assists (season total)")

print(f"\n  Inserted: {inserted2}  |  Skipped: {skipped2}  |  Warnings: {len(warnings2)}")
for w in warnings2:
    print(f"  WARNING: {w}")


# ── Final summary ─────────────────────────────────────────────────────────────
print()
print("=" * 60)
print("DONE")
print(f"  brazukas1 assists inserted  : {inserted}")
print(f"  brazukareceba1 assists inserted: {inserted2}")
total_warnings = warnings + warnings2
if total_warnings:
    print(f"  Total warnings: {len(total_warnings)}")
    for w in total_warnings:
        print(f"    {w}")

"""
import_brazukareceba1.py
Imports historical football data from brazukareceba1.xlsx into Supabase.
Covers Spring 2024 – Winter II 2025 for Brazuka US (team_id=1) and Receba FC (team_id=2).

Sheet layout overview:
  TYPE A — "vertical game blocks" format (most Brazuka 2024-25 sheets):
    col8: date string OR field name OR team/opponent info
    Row pattern (repeating):
      i+0: date string (e.g. "1/7 at 7:40pm")   col8=date_str,  col14='Goals', col15='Assists', col16='Blue Cards'
      i+1: field name                              col8=field_str
      i+2: score row  "TeamA" score1 score2 "TeamB" col8=team_or_opp, col9=score_a, col10=score_b, col11=other_team
      i+3..N: player rows                          col13=player_name, col14=goals, col15=assists, col16=cards
      ...until next date string

  TYPE B — "vertical game blocks" format (Receba Fall 202 sheet):
    col7: date (datetime object) OR opponent name OR score text "NxM vitoria"
    col12: player name, col13: goals, col14: assists, col15: cards
    Row pattern:
      i+0: datetime in col7, col13='Goals', col14='Assists', col15='Blue Cards'
      i+1: opponent name in col7, player rows start
      i+2: score text in col7 "NxM vitoria/derrota/empate"
      i+3..N: more player rows
      ...until next datetime in col7

  TYPE C — "per-row schedule" format (Sheet8 = Brazuka Winter II 2023):
    Same as TYPE A but date is a string in col8 row, with player stats interleaved.
    (Same parser as TYPE A — it handles it.)

  NO-GAME sheets (roster only) — Sheet9, Sheet11, Sheet12, Sheet13,
    Receba FC (Thursday) - Winter I, Receba FC (Thursday) - Spring I,
    Receba FC (Thursday) - Summer 2, Sheet10 (different schedule format, no scores known):
    → These are SKIPPED (no per-game score data to import except Sheet10 which has scores).

  Sheet10 — Receba Winter II 2023 — schedule per row, one game per row starting row 8:
    col6=date, col7=team_name, col8=our_score OR None, col9=opp_score OR None, col10=opponent, col11=field, col12=time
    NOTE: scores missing for some games.

  "Brazuka US (Tuesday) - Spring" — schedule per row, one game per row starting row 9:
    col6=date, col7=away_team, col8=away_score, col9=home_score, col10=home_team, col11=location
"""

import os
import sys
import re
import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl
from supabase import create_client

# ---------------------------------------------------------------------------
# Load environment
# ---------------------------------------------------------------------------
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    sys.exit("ERROR: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# ---------------------------------------------------------------------------
# Player normalizer
# ---------------------------------------------------------------------------
from player_normalizer import PlayerNormalizer
norm = PlayerNormalizer()

# ---------------------------------------------------------------------------
# Load spreadsheet
# ---------------------------------------------------------------------------
XLSX = Path(__file__).parent / "spreadsheet data" / "brazukareceba1.xlsx"
wb = openpyxl.load_workbook(str(XLSX), data_only=True)

# ---------------------------------------------------------------------------
# Counters
# ---------------------------------------------------------------------------
games_inserted = 0
games_skipped = 0
player_warnings = []

# ---------------------------------------------------------------------------
# Season helpers
# ---------------------------------------------------------------------------

def get_or_create_season(name, team_id, start_date, end_date):
    resp = (
        sb.table("seasons")
        .select("id")
        .eq("name", name)
        .eq("team_id", team_id)
        .execute()
    )
    if resp.data:
        return resp.data[0]["id"]
    insert_resp = (
        sb.table("seasons")
        .insert({
            "name": name,
            "team_id": team_id,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
        })
        .execute()
    )
    return insert_resp.data[0]["id"]


# ---------------------------------------------------------------------------
# Game insert helper
# ---------------------------------------------------------------------------

def upsert_game(game_date_str, opponent, home_or_away, result, score_us, score_opp,
                team_id, season_id, venue, field, scorers_known=False):
    """Insert game if not exists. Returns (supabase_id, was_inserted)."""
    global games_inserted, games_skipped
    existing = (
        sb.table("games")
        .select("id")
        .eq("game_date", game_date_str)
        .eq("team_id", team_id)
        .execute()
    )
    if existing.data:
        print(f"  SKIP (already exists): {game_date_str} vs {opponent} (team_id={team_id})")
        games_skipped += 1
        return existing.data[0]["id"], False

    try:
        resp = (
            sb.table("games")
            .insert({
                "game_date": game_date_str,
                "opponent": opponent,
                "home_or_away": home_or_away,
                "result": result,
                "score_brazuka": score_us,
                "score_opponent": score_opp,
                "scorers_known": scorers_known,
                "team_id": team_id,
                "season_id": season_id,
                "venue": venue,
                "field": field,
            })
            .execute()
        )
        new_id = resp.data[0]["id"]
        print(f"  Inserting game {game_date_str} vs {opponent} (team_id={team_id})... OK (id={new_id})")
        games_inserted += 1
        return new_id, True
    except Exception as e:
        if "23505" in str(e) or "duplicate key" in str(e).lower():
            # Date already exists from other spreadsheet — look it up and skip
            existing2 = sb.table("games").select("id").eq("game_date", game_date_str).execute()
            eid = existing2.data[0]["id"] if existing2.data else None
            print(f"  SKIP (date conflict): {game_date_str} vs {opponent} (team_id={team_id})")
            games_skipped += 1
            return eid, False
        raise


def insert_goals_and_appearances(game_id, player_goals, player_assists, player_appearances):
    """
    player_goals: list of (player_name, count)
    player_assists: list of (player_name, count)
    player_appearances: list of player_name
    """
    scorers_known = bool(player_goals or player_assists)

    # Check if goals already exist for this game
    existing_goals = sb.table("goals").select("id").eq("game_id", game_id).execute()
    if existing_goals.data:
        return  # Already imported

    # Build assist lookup: player_name -> count
    assist_lookup = {}
    for pname, cnt in (player_assists or []):
        assist_lookup[pname] = cnt

    for player_name, count in (player_goals or []):
        if not player_name:
            continue
        pid, canonical = norm.resolve_or_flag(player_name)
        if pid is None:
            warning = f"Player not resolved: '{player_name}' (game_id={game_id})"
            player_warnings.append(warning)
            print(f"    WARNING: {warning}")

        sb.table("goals").insert({
            "game_id": game_id,
            "player": player_name,
            "player_id": pid,
            "count": count,
            "notes": None,
            "assist_player": None,
            "assist_player_id": None,
        }).execute()

    for player_name in (player_appearances or []):
        if not player_name:
            continue
        pid, canonical = norm.resolve_or_flag(player_name)
        if pid is None:
            warning = f"Player not resolved (appearance): '{player_name}' (game_id={game_id})"
            player_warnings.append(warning)
            print(f"    WARNING: {warning}")

        existing_app = (
            sb.table("appearances")
            .select("id")
            .eq("game_id", game_id)
            .eq("player", player_name)
            .execute()
        )
        if not existing_app.data:
            sb.table("appearances").insert({
                "game_id": game_id,
                "player": player_name,
                "player_id": pid,
            }).execute()

    if scorers_known:
        sb.table("games").update({"scorers_known": True}).eq("id", game_id).execute()


# ---------------------------------------------------------------------------
# Date parsing helpers
# ---------------------------------------------------------------------------

def parse_date_string(date_str, season_year=None):
    """
    Parse strings like '1/7 at 7:40pm', '1/14 at 6:50pm', '10/1 at 7:45pm'.
    Returns datetime.date or None.
    season_year is used to determine the year (infer from context if needed).
    """
    if not date_str or not isinstance(date_str, str):
        return None
    date_str = date_str.strip()
    # Pattern: M/D at H:MMpm
    m = re.match(r"(\d{1,2})/(\d{1,2})\s+at\s+", date_str, re.IGNORECASE)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        year = season_year or datetime.date.today().year
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    return None


def infer_score(team_a, score_a, score_b, team_b, our_team_names):
    """
    Returns (our_score, opp_score, opponent_name, home_or_away, result).
    our_team_names is a set of strings that identify our team.
    """
    our_names_lower = {n.lower() for n in our_team_names}
    a_lower = (team_a or "").lower().strip()
    b_lower = (team_b or "").lower().strip()

    if a_lower in our_names_lower:
        # We are "team A" — we're home (col8 shows us first)
        return score_a, score_b, team_b, "home", compute_result(score_a, score_b)
    elif b_lower in our_names_lower:
        # We are "team B" — we're away
        return score_b, score_a, team_a, "away", compute_result(score_b, score_a)
    else:
        # Can't determine — log and skip
        return None, None, None, None, None


def compute_result(score_us, score_opp):
    if score_us is None or score_opp is None:
        return None
    if score_us > score_opp:
        return "W"
    elif score_us < score_opp:
        return "L"
    return "D"


def parse_receba_score_text(text):
    """
    Parse score strings like '4x2 vitoria', '6x2 vitoria', '3x2 vitoria',
    '10x8 vitoria', '9x4 vitoria', '11x6 vitória'.
    Returns (our_score, opp_score, result) or (None, None, None).
    """
    if not text or not isinstance(text, str):
        return None, None, None
    text = text.strip().lower()
    m = re.match(r"(\d+)x(\d+)\s*(vitoria|vitória|vitoria |derrota|empate)?", text, re.IGNORECASE)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        keyword = (m.group(3) or "").strip().lower()
        if keyword in ("vitoria", "vitória"):
            return a, b, "W"
        elif keyword == "derrota":
            return a, b, "L"
        elif keyword == "empate":
            return a, b, "D"
        else:
            return a, b, compute_result(a, b)
    return None, None, None


# ---------------------------------------------------------------------------
# TYPE A parser — vertical game blocks in col 8
# ---------------------------------------------------------------------------
OUR_BRAZUKA_NAMES = {"brazuka us", "brazuka"}

def is_date_string(val):
    """Return True if the value looks like a game date string like '1/7 at 7:40pm'."""
    if not isinstance(val, str):
        return False
    return bool(re.match(r"\d{1,2}/\d{1,2}\s+at\s+", val.strip(), re.IGNORECASE))


def parse_type_a_sheet(ws, team_id, league_name, start_date_hint, venue_default="Magnuson Park"):
    """
    Parse a TYPE A Brazuka sheet.
    Returns list of game dicts:
      {date, opponent, home_or_away, result, score_us, score_opp, field,
       goals: [(player, count)], appearances: [player]}
    """
    rows = list(ws.iter_rows(values_only=True))
    games = []
    i = 0
    # Infer year from start_date_hint
    year = start_date_hint.year if start_date_hint else 2024

    while i < len(rows):
        row = list(rows[i])
        col8 = row[8] if len(row) > 8 else None

        if not is_date_string(col8):
            i += 1
            continue

        # Found a game header row
        date_str = col8
        game_date = parse_date_string(date_str, year)
        if game_date is None:
            i += 1
            continue

        # Adjust year: if month seems to wrap (e.g. season starts in Oct, but date is Jan)
        # Trust the hint year; if month < start month - 6, bump year by 1
        if start_date_hint and game_date.month < start_date_hint.month - 3:
            game_date = datetime.date(year + 1, game_date.month, game_date.day)

        # Next row: field name
        field = None
        if i + 1 < len(rows):
            r_next = list(rows[i + 1])
            field_val = r_next[8] if len(r_next) > 8 else None
            if isinstance(field_val, str) and not is_date_string(field_val):
                # Could be field name or already opponent row
                # Field row: only col8 is set in the game area cols 8-11
                score_val = r_next[9] if len(r_next) > 9 else None
                if score_val is None or not isinstance(score_val, (int, float)):
                    field = str(field_val).strip()

        # Score row: look for row with col9 and col10 as numbers and col8/col11 as team names
        score_row_idx = None
        for j in range(i + 1, min(i + 4, len(rows))):
            rj = list(rows[j])
            c8 = rj[8] if len(rj) > 8 else None
            c9 = rj[9] if len(rj) > 9 else None
            c10 = rj[10] if len(rj) > 10 else None
            c11 = rj[11] if len(rj) > 11 else None
            if isinstance(c9, (int, float)) and isinstance(c10, (int, float)):
                score_row_idx = j
                team_a = str(c8).strip() if c8 else ""
                score_a = int(c9)
                score_b = int(c10)
                team_b = str(c11).strip() if c11 else ""
                break

        if score_row_idx is None:
            # No score row found - game might be unplayed, skip
            i += 1
            continue

        # Determine our score/opponent
        our_score, opp_score, opponent, home_or_away, result = infer_score(
            team_a, score_a, score_b, team_b, OUR_BRAZUKA_NAMES
        )

        if opponent is None:
            # Couldn't identify our team - skip
            i += 1
            continue

        # Collect player rows (from score_row_idx+1 until next date string or end)
        player_rows = []
        for k in range(score_row_idx + 1, len(rows)):
            rk = list(rows[k])
            c8k = rk[8] if len(rk) > 8 else None
            if is_date_string(c8k):
                break
            c13 = rk[13] if len(rk) > 13 else None
            c14 = rk[14] if len(rk) > 14 else None  # goals
            c15 = rk[15] if len(rk) > 15 else None  # assists
            # Skip header rows
            if c13 == "Goals" or c14 == "Goals":
                continue
            if c13 and isinstance(c13, str) and c13.strip():
                player_rows.append((c13.strip(), c14, c15))

        goals = []
        appearances = []
        for pname, g_val, a_val in player_rows:
            appearances.append(pname)
            if g_val is not None and isinstance(g_val, (int, float)) and g_val > 0:
                goals.append((pname, int(g_val)))

        games.append({
            "date": game_date,
            "opponent": opponent,
            "home_or_away": home_or_away,
            "result": result,
            "score_us": our_score,
            "score_opp": opp_score,
            "field": field,
            "goals": goals,
            "appearances": appearances,
        })

        i = score_row_idx + 1

    return games


# ---------------------------------------------------------------------------
# TYPE B parser — Receba Fall 202 style (col 7 game blocks)
# ---------------------------------------------------------------------------

def parse_type_b_sheet(ws, team_id, league_name, start_date_hint):
    """
    Parse a TYPE B Receba sheet where col7 holds game info.
    Game block starts when col7 is a datetime object.
    Score comes as text "NxM vitoria" in col7.
    Player data: col12=name, col13=goals, col14=assists.
    """
    rows = list(ws.iter_rows(values_only=True))
    games = []
    i = 0

    while i < len(rows):
        row = list(rows[i])
        col7 = row[7] if len(row) > 7 else None

        if not isinstance(col7, (datetime.datetime, datetime.date)):
            i += 1
            continue

        # Found a game date
        game_date = col7.date() if isinstance(col7, datetime.datetime) else col7

        # Collect following rows until next datetime in col7
        opponent = None
        score_text = None
        our_score = None
        opp_score = None
        result = None
        player_rows = []

        for j in range(i + 1, len(rows)):
            rj = list(rows[j])
            c7 = rj[7] if len(rj) > 7 else None

            if isinstance(c7, (datetime.datetime, datetime.date)):
                # Next game block
                break

            if isinstance(c7, str) and c7.strip():
                val = c7.strip()
                # Is it a score text?
                if re.match(r"\d+x\d+", val, re.IGNORECASE):
                    score_text = val
                    s_us, s_opp, res = parse_receba_score_text(val)
                    our_score = s_us
                    opp_score = s_opp
                    result = res
                elif opponent is None:
                    opponent = val

            # Player data in col12/13/14
            c12 = rj[12] if len(rj) > 12 else None
            c13 = rj[13] if len(rj) > 13 else None  # goals
            if c12 and isinstance(c12, str) and c12.strip() and c12.strip() != "Goals":
                player_rows.append((c12.strip(), c13))

        if opponent is None:
            i += 1
            continue

        goals = []
        appearances = []
        for pname, g_val in player_rows:
            appearances.append(pname)
            if g_val is not None and isinstance(g_val, (int, float)) and g_val > 0:
                goals.append((pname, int(g_val)))

        games.append({
            "date": game_date,
            "opponent": opponent,
            "home_or_away": "away",  # Receba plays away (Redmond)
            "result": result,
            "score_us": our_score,
            "score_opp": opp_score,
            "field": None,
            "goals": goals,
            "appearances": appearances,
        })

        i += 1

    return games


# ---------------------------------------------------------------------------
# TYPE C parser — "Brazuka Spring" per-row schedule (1 game per row)
# ---------------------------------------------------------------------------

def parse_brazuka_spring_sheet(ws, team_id, league_name, start_date_hint, venue_default="Magnuson Park"):
    """
    col6=date(datetime), col7=away_team, col8=away_score, col9=home_score, col10=home_team, col11=location
    Rows start at index 9 (after headers).
    No per-player goal data available.
    """
    rows = list(ws.iter_rows(values_only=True))
    games = []
    seen_dates = set()

    for row in rows[9:]:
        r = list(row)
        if len(r) < 11:
            continue
        date_val = r[6]
        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            continue
        game_date = date_val.date() if isinstance(date_val, datetime.datetime) else date_val
        if game_date in seen_dates:
            continue
        seen_dates.add(game_date)

        away_team = str(r[7]).strip() if r[7] else ""
        away_score = int(r[8]) if r[8] is not None and isinstance(r[8], (int, float)) else None
        home_score = int(r[9]) if r[9] is not None and isinstance(r[9], (int, float)) else None
        home_team = str(r[10]).strip() if r[10] else ""
        location = str(r[11]).strip() if r[11] else None

        if away_score is None and home_score is None:
            continue  # Skip unplayed games

        our_score, opp_score, opponent, home_or_away, result = infer_score(
            away_team, away_score, home_score, home_team, OUR_BRAZUKA_NAMES
        )
        if opponent is None:
            continue

        games.append({
            "date": game_date,
            "opponent": opponent,
            "home_or_away": home_or_away,
            "result": result,
            "score_us": our_score,
            "score_opp": opp_score,
            "field": location,
            "goals": [],
            "appearances": [],
        })

    return games


# ---------------------------------------------------------------------------
# TYPE D parser — Sheet8 (Brazuka Winter II 2023) - same as TYPE A
# ---------------------------------------------------------------------------
# Sheet8 uses the same vertical-block format as TYPE A, just without a Start Date row.
# The start date can be inferred from the first game date.


# ---------------------------------------------------------------------------
# TYPE E parser — Sheet10 (Receba Winter II 2023)
# Each row from row 8 onward is one game:
# col6=date, col7=team_name("Receba FC"), col8=our_score OR None,
# col9=opp_score OR None, col10=opponent, col11=field, col12=time
# ---------------------------------------------------------------------------

def parse_receba_schedule_sheet(ws, team_id, league_name, start_date_hint, venue_default="Redmond"):
    """Parse Sheet10 style: one game per row."""
    rows = list(ws.iter_rows(values_only=True))
    games = []

    for row in rows[8:]:
        r = list(row)
        if len(r) < 11:
            continue
        date_val = r[6]
        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            continue
        game_date = date_val.date() if isinstance(date_val, datetime.datetime) else date_val
        our_score = int(r[8]) if r[8] is not None and isinstance(r[8], (int, float)) else None
        opp_score = int(r[9]) if r[9] is not None and isinstance(r[9], (int, float)) else None
        opponent = str(r[10]).strip() if r[10] else None
        field = str(r[11]).strip() if r[11] else None

        if opponent is None:
            continue
        if our_score is None and opp_score is None:
            continue

        result = compute_result(our_score, opp_score)

        games.append({
            "date": game_date,
            "opponent": opponent,
            "home_or_away": "away",
            "result": result,
            "score_us": our_score,
            "score_opp": opp_score,
            "field": field,
            "goals": [],
            "appearances": [],
        })

    return games


# ---------------------------------------------------------------------------
# Metadata extractor helper
# ---------------------------------------------------------------------------

def get_sheet_metadata(ws):
    """Extract league name and start date from rows 0-5."""
    rows = list(ws.iter_rows(values_only=True, min_row=1, max_row=7))
    league = None
    start_date = None
    weekday = None
    location = None
    for row in rows:
        r = list(row)
        if r[0] == "League" and r[1]:
            league = str(r[1]).strip()
        elif r[0] == "Start Date" and r[1]:
            if isinstance(r[1], (datetime.datetime, datetime.date)):
                start_date = r[1].date() if isinstance(r[1], datetime.datetime) else r[1]
        elif r[0] == "Weekday" and r[1]:
            weekday = str(r[1]).strip()
        elif r[0] == "Location" and r[1]:
            location = str(r[1]).strip()
    return league, start_date, weekday, location


# ---------------------------------------------------------------------------
# Sheet definitions
# ---------------------------------------------------------------------------

# (sheet_name, team_id, parser_type)
# team_id: 1=Brazuka US, 2=Receba FC
# parser_type: A, B, C, D, E, or SKIP

SHEET_CONFIG = [
    # Brazuka US sheets with TYPE A vertical-block game data
    ("Brazuka US (Tuesday) - Winter I", 1, "A"),   # 2025 Winter II
    ("Sheet7",                          1, "A"),   # 2024 Winter I
    ("Brazuka US (Tuesday) - Fall 202", 1, "A"),   # 2024 Fall
    ("Brazuka US (Tuesday) - Summer 2", 1, "A"),   # 2024 Summer
    ("Brazuka US (Tuesday) - Spring 2", 1, "A"),   # 2024 Spring
    ("Sheet8",                          1, "D"),   # 2023 Winter II (TYPE A compatible)
    # Brazuka US sheets with per-row schedule (no scorer data)
    ("Brazuka US (Tuesday) - Spring",   1, "C"),   # 2023 Spring
    # Brazuka sheets with roster only - SKIP
    ("Sheet11",                         1, "SKIP"),  # 2023 Summer
    ("Sheet12",                         1, "SKIP"),  # 2023 Fall
    ("Sheet13",                         1, "SKIP"),  # 2023 Winter I
    # Receba FC sheets with TYPE B game data
    ("Receba FC (Thursday) - Fall 202", 2, "B"),   # 2024 Fall
    # Receba FC sheet with per-row schedule
    ("Sheet10",                         2, "E"),   # 2023 Winter II
    # Receba FC sheets with roster only - SKIP
    ("Sheet9",                          2, "SKIP"),  # 2023 Winter I
    ("Receba FC (Thursday) - Winter I", 2, "SKIP"),  # 2024 Winter I
    ("Receba FC (Thursday) - Spring I", 2, "SKIP"),  # 2024 Spring
    ("Receba FC (Thursday) - Summer 2", 2, "SKIP"),  # 2024 Summer
]

# Season_id cache: (season_name, team_id) -> id
season_id_cache = {}


def get_season_id(season_name, team_id, game_dates):
    key = (season_name, team_id)
    if key in season_id_cache:
        return season_id_cache[key]
    dates = sorted([d for d in game_dates if d])
    start_dt = dates[0] if dates else None
    end_dt = dates[-1] if dates else None
    sid = get_or_create_season(season_name, team_id, start_dt, end_dt)
    season_id_cache[key] = sid
    return sid


# ---------------------------------------------------------------------------
# Main processing loop
# ---------------------------------------------------------------------------

VENUE_MAP = {
    "Magnuson": "Magnuson Park",
    "Redmond": "Redmond",
    "SODO": "SODO",
}

print("=== Processing brazukareceba1.xlsx ===\n")

for sheet_name, team_id, parser_type in SHEET_CONFIG:
    if parser_type == "SKIP":
        print(f"--- Skipping {sheet_name} (no game data) ---")
        continue

    if sheet_name not in wb.sheetnames:
        print(f"--- Sheet '{sheet_name}' not found, skipping ---")
        continue

    ws = wb[sheet_name]
    league_name, start_date, weekday, location = get_sheet_metadata(ws)

    venue_default = VENUE_MAP.get(location, location or "Unknown")
    team_label = "Brazuka US" if team_id == 1 else "Receba FC"

    print(f"--- Processing '{sheet_name}' | {team_label} | League: {league_name} | Start: {start_date} ---")

    if parser_type == "A" or parser_type == "D":
        if start_date is None and parser_type == "D":
            # Sheet8: start date is in row 5
            rows_raw = list(ws.iter_rows(values_only=True))
            if len(rows_raw) > 5 and rows_raw[5][1]:
                v = rows_raw[5][1]
                start_date = v.date() if isinstance(v, datetime.datetime) else v
        games = parse_type_a_sheet(ws, team_id, league_name, start_date, venue_default)
    elif parser_type == "B":
        games = parse_type_b_sheet(ws, team_id, league_name, start_date)
    elif parser_type == "C":
        games = parse_brazuka_spring_sheet(ws, team_id, league_name, start_date, venue_default)
    elif parser_type == "E":
        games = parse_receba_schedule_sheet(ws, team_id, league_name, start_date, venue_default)
    else:
        print(f"  Unknown parser type {parser_type}, skipping")
        continue

    if not games:
        print(f"  No games parsed from {sheet_name}")
        continue

    # Build season_id
    all_dates = [g["date"] for g in games]
    season_id = get_season_id(league_name or "Unknown", team_id, all_dates)
    print(f"  Season id={season_id}, {len(games)} games parsed")

    for g in games:
        game_date_str = g["date"].isoformat()
        game_id, was_inserted = upsert_game(
            game_date_str=game_date_str,
            opponent=g["opponent"],
            home_or_away=g["home_or_away"],
            result=g["result"],
            score_us=g["score_us"],
            score_opp=g["score_opp"],
            team_id=team_id,
            season_id=season_id,
            venue=venue_default,
            field=g.get("field"),
            scorers_known=bool(g["goals"]),
        )

        if was_inserted and (g["goals"] or g["appearances"]):
            insert_goals_and_appearances(game_id, g["goals"], [], g["appearances"])

    print()

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
print("=== Summary ===")
print(f"  Games inserted  : {games_inserted}")
print(f"  Games skipped   : {games_skipped}")
print(f"  Player warnings : {len(player_warnings)}")
for w in player_warnings:
    print(f"    WARNING: {w}")
